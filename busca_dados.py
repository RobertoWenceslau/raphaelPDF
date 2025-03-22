import pdfplumber
import pandas as pd
from datetime import datetime
import tkinter as tk
from tkinter import filedialog
import re
from openpyxl.workbook import Workbook
from openpyxl import load_workbook


# Carregar a tabela de CIDs
cid_mapping = pd.read_csv("cid_mapping.csv", sep=";", encoding="ISO-8859-1")
cid_dict = dict(zip(cid_mapping["SUBCAT"], cid_mapping["DESCRICAO"]))


def corrigir_cid(cid):
    """
    Corrige possíveis erros no formato do CID.
    Exemplo: "1200" → "I200".
    Retorna o CID corrigido e uma mensagem de correção (se houver).
    """
    cid_original = cid
    if len(cid) == 4 and cid[0].isdigit():
        # Se o primeiro caractere for um número, substitua por "I"
        cid = "I" + cid[1:]
        mensagem_correcao = f"CID: {cid_original} → {cid}"
    else:
        mensagem_correcao = ""
    return cid, mensagem_correcao


def extrair_data(texto, padrao_busca, linhas_contexto=5):
    """
    Extrai uma data de um texto baseado em um padrão de busca.
    Procura por padrões de data DD/MM/YY ou DD/MM/YYYY próximos ao padrão de busca.
    """
    # Padrão regex para datas no formato DD/MM/YY ou DD/MM/YYYY
    padrao_data = r'\b(\d{2}/\d{2}/\d{2,4})\b'

    # Procurar o padrão de busca no texto
    indice = texto.find(padrao_busca)
    if indice == -1:
        return None

    # Extrair um trecho do texto ao redor do padrão de busca
    inicio = max(0, indice - 1) # 50
    fim = min(len(texto), indice + 100) # 100
    trecho = texto[inicio:fim]


    # Procurar datas no trecho
    datas = re.findall(padrao_data, trecho)
    if datas:
        # Retornar a primeira data encontrada, formatada como DD/MM/YY
        data = datas[0]
        try:
            if len(data) == 10:  # DD/MM/YYYY
                data_obj = datetime.strptime(data, "%d/%m/%Y")
            else:  # DD/MM/YY
                data_obj = datetime.strptime(data, "%d/%m/%y")
            return data_obj.strftime("%d/%m/%y")
        except ValueError:
            return None

    # Se não encontrar datas no trecho, procurar nas próximas linhas
    linhas = texto.split('\n')
    for i, linha in enumerate(linhas):
        if padrao_busca in linha:
            # Verificar as próximas linhas_contexto linhas
            for j in range(i, min(i + linhas_contexto, len(linhas))):
                datas = re.findall(padrao_data, linhas[j])
                if datas:
                    data = datas[0]
                    try:
                        if len(data) == 10:  # DD/MM/YYYY
                            data_obj = datetime.strptime(data, "%d/%m/%Y")
                        else:  # DD/MM/YY
                            data_obj = datetime.strptime(data, "%d/%m/%y")
                        return data_obj.strftime("%d/%m/%y")
                    except ValueError:
                        continue

    return None


def extrair_dados(pdf_path):
    dados = []
    with pdfplumber.open(pdf_path) as pdf:
        # Processar apenas a primeira página
        page = pdf.pages[0]
        texto = page.extract_text()

        # Verificar se o texto foi extraído corretamente
        if texto is None:
            return pd.DataFrame(dados)

        # Inicializar variáveis com valores padrão
        beneficiario = "Beneficiário não encontrado"
        prestador = "UNIMED FLORIANOPOLIS"  # Valor padrão
        cid = "CID não encontrado"
        motivo = "MOTIVO NÃO INFORMADO"
        data_admissao = "Data de admissão não encontrada"
        data_alta = "Data de alta não encontrada"
        tipo_internacao = "CLINICA"
        carater = "ELETIVO"  # Valor padrão

        # Extrair dados do beneficiário
        if "Beneficiário(a)" in texto:
            beneficiario_parte = texto.split("Beneficiário(a)")[1].split("\n")[0].strip()
            # Remover o número do beneficiário, se presente
            partes = beneficiario_parte.split()

            if len(partes) > 1:
                beneficiario = " ".join(partes[1:])  # Remove o primeiro item (número)


        # Extrair nome do prestador
        if "Executora" in texto:
            prestador_pdf = texto.split("Executora")[1].split("\n")[0].strip()
            if prestador_pdf:
                prestador = "UNIMED FLORIANOPOLIS"

        # Extrair CID Principal
        if "Principal" in texto:
            cid_parte = texto.split("Principal")[1].split("\n")[0].strip()

            if cid_parte:
                cid = cid_parte
                cid, mensagem_correcao = corrigir_cid(cid)  # Corrigir o CID, se necessário
                if cid in cid_dict:
                    motivo = cid_dict[cid]


        # Extrair datas de internação e alta
        data_admissao_extraida = extrair_data(texto, "Atendimento ")
        if data_admissao_extraida:
            data_admissao = data_admissao_extraida

        data_alta_extraida = extrair_data(texto, "Alta ")
        if data_alta_extraida:
            data_alta = data_alta_extraida

        # Extrair todas as diárias e datas de acomodação
        diarias = []
        for line in texto.split("\n"):
            if "DIÁRIA DE" in line:
                partes = line.split()
                if len(partes) >= 3:
                    tipo_diaria = " ".join([p for p in partes[2:] if not re.match(r'\d{2}/\d{2}/\d{2,4}', p)])
                    # Extrair a data da diária
                    datas_na_linha = re.findall(r'\b(\d{2}/\d{2}/\d{2,4})\b', line)
                    if datas_na_linha:
                        data_diaria = datas_na_linha[0]
                        diarias.append({"tipo": tipo_diaria, "data": data_diaria})

        # Determinar o CARÁTER com base nas diárias
        if diarias:
            diarias_com_datas = []
            for diaria in diarias:
                try:
                    # Padronizar formato de data
                    if len(diaria["data"]) == 10:  # DD/MM/YYYY
                        data_obj = datetime.strptime(diaria["data"], "%d/%m/%Y")
                    else:  # DD/MM/YY
                        data_obj = datetime.strptime(diaria["data"], "%d/%m/%y")
                    diarias_com_datas.append({"tipo": diaria["tipo"], "data": data_obj})
                except ValueError:
                    continue

            # Verificar se há alguma UTI no registro de diárias
            tem_uti = any("UTI" in d["tipo"].upper() for d in diarias_com_datas)

            # Se houver UTI ou palavras-chave de urgência, é URGÊNCIA
            if tem_uti or "URGENTE" in texto.upper() or "EMERGÊNCIA" in texto.upper():
                carater = "URGENCIA"

        # Inferir TIPO DE INTERNAÇÃO com base nos procedimentos
        procedimentos_cirurgicos = ["CATETERISMO", "CIRURGIA", "RESSECÇÃO", "TAXA DE SALA CIRÚRGICA",
                                    "CIRURGICO", "CIRÚRGICO", "OPERAÇÃO", "PROCEDIMENTO CIRÚRGICO"]
        if any(proc in texto.upper() for proc in procedimentos_cirurgicos):
            tipo_internacao = "CIRURGICA"
        else:
            tipo_internacao = "CLINICA"

        # Extrair motivo da internação de outras partes do documento se não foi encontrado pelo CID
        if motivo == "MOTIVO NÃO INFORMADO":
            palavras_chave_motivo = ["DIAGNÓSTICO", "DIAGNÓSTICO PRINCIPAL", "MOTIVO INTERNAÇÃO",
                                     "MOTIVO DA INTERNAÇÃO", "CAUSA"]
            for palavra in palavras_chave_motivo:
                if palavra in texto.upper():
                    try:
                        trecho = texto.split(palavra)[1].split("\n")[0].strip()
                        if len(trecho) > 5 and not trecho.startswith("Data") and not trecho.isdigit():
                            motivo = trecho
                            break
                    except:
                        continue

        # Adicionar dados à lista
        dados.append({
            "UF": "SC",
            "TIPO DE INTERNAÇÃO": tipo_internacao,
            "CARÁTER": carater,
            "Beneficiário Atendido": beneficiario,
            "ACOMODAÇÃO": "APARTAMENTO",  # Acomodação fixa
            "PRESTADOR": prestador,
            "CNPJ": "77.658.611/0001-08",
            "DATA DA ADMISSÃO": data_admissao,
            "DATA DA ALTA": data_alta,
            "MOTIVO DA INTERNAÇÃO": motivo
        })

    # Remover duplicatas
    df = pd.DataFrame(dados)
    if not df.empty:
        df = df.drop_duplicates()

    return df


def selecionar_arquivo():
    root = tk.Tk()
    root.withdraw()  # Ocultar Janela principal
    caminho_arquivo = filedialog.askopenfilename(title="Selecione o arquivo PDF",
                                                 filetypes=(("Arquivos PDF", ".pdf"), ("Todos os Arquivos", ".*")))
    return caminho_arquivo


def dados_pdf(pdf_path, primeira_iteracao):
    # pdf_path = selecionar_arquivo()
    tabela = extrair_dados(pdf_path)
    if not tabela.empty:
        excel_path = "censo_internados.xlsx"

        if primeira_iteracao:
            # Na primeira iteração, cria ou sobrescreve o arquivo
            tabela.to_excel(excel_path, index=False)
            print("\n\033[34mArquivo criado/sobrescrito com sucesso na primeira iteração.\033[m")
        else:
            try:
                # Nas iterações subsequentes, adiciona os dados
                workbook = load_workbook(excel_path)
                sheet = workbook.active

                # Calcula a próxima linha vazia
                next_row = sheet.max_row + 1

                # Adiciona os dados a partir da próxima linha
                for row in tabela.itertuples(index=False, name=None):
                    sheet.append(row)

                workbook.save(excel_path)
                print("Dados adicionados com sucesso na iteração subsequente.")
            except FileNotFoundError:
                print(f"Erro: o arquivo {excel_path} não foi encontrado.")
    else:
        print("Nenhum dado foi extraído do PDF.")


# Exemplo de uso
