from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def formatar_excel(arquivo_excel):
    # Carrega o arquivo Excel gerado
    workbook = load_workbook(arquivo_excel)
    sheet = workbook.active

    # Ajusta a largura de cada coluna para caber no conteúdo
    for col in sheet.columns:
        max_tamanho = 0
        col_letter = get_column_letter(col[0].column)  # Obtém a letra da coluna

        # Determina o tamanho máximo do conteúdo da coluna
        for cell in col:
            if cell.value:
                max_tamanho = max(max_tamanho, len(str(cell.value)))

        # Ajusta a largura da coluna com base no tamanho máximo
        sheet.column_dimensions[col_letter].width = max_tamanho + 2

    # Salva as alterações no arquivo Excel
    workbook.save(arquivo_excel)
    print(f"\n\033[33mO arquivo '{arquivo_excel}' foi formatado com sucesso.\033[m")

